#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Create test Excel files with Array Formulas and External Links for testing
"""

import os
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula

def create_test_excel():
    """Create a test Excel file with array formulas"""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    # Add some regular data
    ws['A1'] = 'Data1'
    ws['B1'] = 'Data2'
    ws['A2'] = 10
    ws['B2'] = 20
    
    # Add array formula
    # Note: Array formulas in openpyxl are handled differently
    ws['C1'] = '=SUM(A1:B2)'
    
    # Add some external reference like formulas (simulated)
    ws['D1'] = '=[1]Sheet1!A1'  # This simulates an external reference
    ws['E1'] = '=[2]Table!B1'   # This simulates another external reference
    
    save_path = '/home/runner/work/watchdog/watchdog/test_setup/test_file.xlsx'
    wb.save(save_path)
    print(f"Test Excel file created at: {save_path}")
    return save_path

def create_test_excel_with_external_links():
    """Create a test Excel file with external links"""
    wb = Workbook()
    ws = wb.active
    ws.title = "MainSheet"
    
    # Add formulas with external references
    ws['A1'] = '=[1]Sheet1!A1'
    ws['B1'] = '=[2]Data!B1'
    ws['C1'] = '=[3]Table!C1'
    
    save_path = '/home/runner/work/watchdog/watchdog/test_setup/test_external_links.xlsx'
    wb.save(save_path)
    print(f"Test Excel file with external links created at: {save_path}")
    return save_path

if __name__ == "__main__":
    create_test_excel()
    create_test_excel_with_external_links()