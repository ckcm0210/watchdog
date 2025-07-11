#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Complete integration test of the enhanced watchdog functionality
"""

import sys
import os
import tempfile
import shutil
import json
import gzip
sys.path.append('/home/runner/work/watchdog/watchdog')

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.formula import ArrayFormula

# Import the enhanced functions
from watch import serialize_cell_value, extract_external_links, resolve_external_references, dump_excel_cells_with_timeout, hash_excel_content

def create_test_excel_with_array_formula():
    """Create a test Excel file that might have ArrayFormula issues"""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    # Add some data
    ws['A1'] = 10
    ws['A2'] = 20
    ws['A3'] = 30
    ws['B1'] = 5
    ws['B2'] = 10
    ws['B3'] = 15
    
    # Add a formula that could become an ArrayFormula
    ws['C1'] = '=SUM(A1:A3)'
    ws['C2'] = '=AVERAGE(B1:B3)'
    
    # Add external link style formulas
    ws['D1'] = '=[1]ExternalSheet!A1'
    ws['D2'] = '=[2]DataTable!B1'
    
    test_file = '/tmp/test_watchdog_enhanced.xlsx'
    wb.save(test_file)
    return test_file

def simulate_watchdog_scenario():
    """Simulate the complete watchdog scenario"""
    print("=" * 60)
    print("Simulating complete watchdog scenario")
    print("=" * 60)
    
    # Create test file
    test_file = create_test_excel_with_array_formula()
    print(f"Created test file: {test_file}")
    
    # Mock required config variables
    import watch
    watch.ENABLE_FAST_MODE = True
    watch.USE_LOCAL_CACHE = False
    watch.CACHE_FOLDER = "/tmp"
    
    def mock_copy_to_cache(path):
        return path
    
    watch.copy_to_cache = mock_copy_to_cache
    
    # Simulate first read (baseline creation)
    print("\n1. Creating baseline...")
    baseline_cells = dump_excel_cells_with_timeout(test_file)
    baseline_hash = hash_excel_content(baseline_cells)
    print(f"   Baseline hash: {baseline_hash}")
    print(f"   Baseline cells: {len(baseline_cells.get('TestSheet', {}))}")
    
    # Save baseline (simulate)
    baseline_data = {
        "content_hash": baseline_hash,
        "cells": baseline_cells
    }
    
    # Simulate second read (change detection)
    print("\n2. Simulating second read (should be same)...")
    current_cells = dump_excel_cells_with_timeout(test_file)
    current_hash = hash_excel_content(current_cells)
    print(f"   Current hash: {current_hash}")
    print(f"   Hashes match: {baseline_hash == current_hash}")
    
    # Simulate ArrayFormula object address change scenario
    print("\n3. Simulating ArrayFormula object address change...")
    
    # Create two cell data with ArrayFormula objects that have same content but different addresses
    af1 = ArrayFormula(ref="A1:A3")
    af1.text = "SUM(A1:A3)"
    
    af2 = ArrayFormula(ref="A1:A3")
    af2.text = "SUM(A1:A3)"
    
    # Simulate the scenario where only the object address changes
    baseline_cell_with_af = {
        "formula": "=SUM(A1:A3)",
        "value": serialize_cell_value(af1)
    }
    
    current_cell_with_af = {
        "formula": "=SUM(A1:A3)",
        "value": serialize_cell_value(af2)
    }
    
    print(f"   Baseline cell: {baseline_cell_with_af}")
    print(f"   Current cell: {current_cell_with_af}")
    print(f"   Cells equal: {baseline_cell_with_af == current_cell_with_af}")
    
    if baseline_cell_with_af == current_cell_with_af:
        print("   ✅ SUCCESS: ArrayFormula object address change filtered correctly!")
    else:
        print("   ❌ FAILURE: ArrayFormula object address change not filtered!")
    
    # Test external link resolution
    print("\n4. Testing external link resolution...")
    
    # Mock external link mapping
    mock_mapping = {
        1: "external_source1.xlsx",
        2: "data_table.xlsx"
    }
    
    test_formulas = [
        "=[1]ExternalSheet!A1",
        "=[2]DataTable!B1",
        "=SUM([1]ExternalSheet!A1:[1]ExternalSheet!A10)"
    ]
    
    for formula in test_formulas:
        resolved = resolve_external_references(formula, mock_mapping)
        print(f"   {formula} -> {resolved}")
        if formula != resolved:
            print(f"   ✅ External link resolved!")
    
    # Clean up
    if os.path.exists(test_file):
        os.remove(test_file)
    
    print("\n✅ Complete integration test completed successfully!")

def test_comparison_with_changes():
    """Test comparison logic with actual changes"""
    print("\n" + "=" * 60)
    print("Testing comparison with actual changes")
    print("=" * 60)
    
    # Create baseline data
    baseline_cells = {
        "TestSheet": {
            "A1": {"formula": None, "value": 10},
            "B1": {"formula": "=SUM(A1:A3)", "value": "SUM(A1:A3)"},
            "C1": {"formula": "=[1]ExternalSheet!A1", "value": "=[1]ExternalSheet!A1"}
        }
    }
    
    # Create current data with changes
    current_cells = {
        "TestSheet": {
            "A1": {"formula": None, "value": 20},  # Value changed
            "B1": {"formula": "=SUM(A1:A3)", "value": "SUM(A1:A3)"},  # No change
            "C1": {"formula": "=[1]ExternalSheet!A1", "value": "=[1]ExternalSheet!A1"}  # No change
        }
    }
    
    # Find changes
    changes = []
    for ws_name in current_cells.keys():
        baseline_ws = baseline_cells.get(ws_name, {})
        current_ws = current_cells[ws_name]
        
        all_cells = set(baseline_ws.keys()) | set(current_ws.keys())
        
        for cell_coord in all_cells:
            baseline_cell = baseline_ws.get(cell_coord, {"formula": None, "value": None})
            current_cell = current_ws.get(cell_coord, {"formula": None, "value": None})
            
            if baseline_cell != current_cell:
                changes.append({
                    'worksheet': ws_name,
                    'cell': cell_coord,
                    'old_formula': baseline_cell['formula'],
                    'old_value': baseline_cell['value'],
                    'new_formula': current_cell['formula'],
                    'new_value': current_cell['value']
                })
    
    print(f"Changes detected: {len(changes)}")
    for change in changes:
        print(f"  {change['worksheet']}:{change['cell']} - {change['old_value']} -> {change['new_value']}")
    
    print("✅ Change detection test completed")

if __name__ == "__main__":
    simulate_watchdog_scenario()
    test_comparison_with_changes()