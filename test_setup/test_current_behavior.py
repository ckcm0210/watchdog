#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test script to understand ArrayFormula handling and external link parsing
"""

import sys
import os
sys.path.append('/home/runner/work/watchdog/watchdog')

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
import zipfile
import xml.etree.ElementTree as ET

def test_array_formula_handling():
    """Test how ArrayFormula objects are handled"""
    print("=" * 60)
    print("Testing ArrayFormula handling")
    print("=" * 60)
    
    wb = load_workbook('/home/runner/work/watchdog/watchdog/test_setup/test_file.xlsx')
    ws = wb.active
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                print(f"Cell {cell.coordinate}:")
                print(f"  Value: {cell.value}")
                print(f"  Type: {type(cell.value)}")
                print(f"  Data type: {cell.data_type}")
                if hasattr(cell.value, 'formula'):
                    print(f"  Formula: {cell.value.formula}")
                if isinstance(cell.value, ArrayFormula):
                    print(f"  ArrayFormula detected!")
                    print(f"    Formula text: {cell.value.formula}")
                    print(f"    Object repr: {repr(cell.value)}")
                print()
    
    wb.close()

def test_external_link_parsing():
    """Test parsing external links from Excel file"""
    print("=" * 60)
    print("Testing external link parsing")
    print("=" * 60)
    
    file_path = '/home/runner/work/watchdog/watchdog/test_setup/test_external_links.xlsx'
    
    # First, check what formulas we have
    wb = load_workbook(file_path)
    ws = wb.active
    
    print("Formulas in the file:")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.data_type == 'f':
                print(f"Cell {cell.coordinate}: {cell.value}")
    
    wb.close()
    
    # Now try to parse as zip file
    print("\nParsing as zip file:")
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            # List all files
            print("Files in the Excel zip:")
            for name in zip_ref.namelist():
                print(f"  {name}")
            
            # Look for external links
            external_links = [name for name in zip_ref.namelist() if 'externalLink' in name.lower()]
            print(f"\nFound external link files: {external_links}")
            
            # Look for relationships
            rels_files = [name for name in zip_ref.namelist() if '_rels' in name and name.endswith('.rels')]
            print(f"Found relationship files: {rels_files}")
            
            # Try to read workbook.xml.rels
            workbook_rels = None
            for rels_file in rels_files:
                if 'workbook.xml.rels' in rels_file:
                    workbook_rels = rels_file
                    break
            
            if workbook_rels:
                print(f"\nReading {workbook_rels}:")
                content = zip_ref.read(workbook_rels)
                print(content.decode('utf-8'))
                
    except Exception as e:
        print(f"Error parsing zip file: {e}")

if __name__ == "__main__":
    test_array_formula_handling()
    test_external_link_parsing()