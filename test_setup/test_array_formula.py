#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test ArrayFormula handling and create example scenarios
"""

import sys
import os
import tempfile
import shutil
sys.path.append('/home/runner/work/watchdog/watchdog')

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.formula import ArrayFormula
import json

def create_array_formula_test():
    """Create test files with ArrayFormula to understand the issue"""
    
    # Create a workbook with array formula
    wb = Workbook()
    ws = wb.active
    
    # Add some data
    ws['A1'] = 10
    ws['A2'] = 20
    ws['A3'] = 30
    ws['B1'] = 5
    ws['B2'] = 15
    ws['B3'] = 25
    
    # Try to create an array formula
    # Note: openpyxl may not create ArrayFormula objects directly
    # but we can test with regular formulas that might become ArrayFormulas
    
    # This might become an ArrayFormula in certain Excel versions
    ws['C1'] = '=SUM(A1:A3*B1:B3)'
    
    # Save the file
    test_file = '/tmp/test_array_formula.xlsx'
    wb.save(test_file)
    
    print("Created test file with potential array formula")
    
    # Now read it back and see how it's handled
    wb2 = load_workbook(test_file)
    ws2 = wb2.active
    
    print("\nReading back the file:")
    for row in ws2.iter_rows():
        for cell in row:
            if cell.value is not None:
                print(f"Cell {cell.coordinate}:")
                print(f"  Value: {cell.value}")
                print(f"  Type: {type(cell.value)}")
                print(f"  Type name: {type(cell.value).__name__}")
                print(f"  Data type: {cell.data_type}")
                
                # Check if it's an ArrayFormula
                if isinstance(cell.value, ArrayFormula):
                    print(f"  ArrayFormula detected!")
                    print(f"    Formula: {cell.value.formula}")
                    print(f"    Repr: {repr(cell.value)}")
                    
                    # Test two different ArrayFormula objects with same formula
                    af1 = ArrayFormula(formula=cell.value.formula)
                    af2 = ArrayFormula(formula=cell.value.formula)
                    print(f"    AF1 repr: {repr(af1)}")
                    print(f"    AF2 repr: {repr(af2)}")
                    print(f"    AF1 == AF2: {af1 == af2}")
                    print(f"    AF1.formula == AF2.formula: {af1.formula == af2.formula}")
                    print(f"    repr(AF1) == repr(AF2): {repr(af1) == repr(af2)}")
                
                print()
    
    wb2.close()
    
    # Clean up
    if os.path.exists(test_file):
        os.remove(test_file)

def test_serialize_comparison():
    """Test the serialize functions from different versions"""
    
    print("=" * 60)
    print("Testing serialize functions")
    print("=" * 60)
    
    # Create some test ArrayFormula objects
    af1 = ArrayFormula(formula="SUM(A1:A3)")
    af2 = ArrayFormula(formula="SUM(A1:A3)")  # Same formula
    af3 = ArrayFormula(formula="SUM(B1:B3)")  # Different formula
    
    print(f"af1 formula: {af1.formula}")
    print(f"af2 formula: {af2.formula}")
    print(f"af3 formula: {af3.formula}")
    print(f"af1 == af2: {af1 == af2}")
    print(f"af1.formula == af2.formula: {af1.formula == af2.formula}")
    print(f"repr(af1): {repr(af1)}")
    print(f"repr(af2): {repr(af2)}")
    print(f"repr(af1) == repr(af2): {repr(af1) == repr(af2)}")
    
    # Test with current serialize_cell_value (from watch.py)
    def current_serialize(value):
        if value is None:
            return None
        elif isinstance(value, (int, float, str, bool)):
            return value
        else:
            return str(value)
    
    # Test with v2 serialize_cell_value
    def v2_serialize(value):
        if value is None:
            return None
        if type(value).__name__ == "ArrayFormula":
            return str(value.formula)
        if hasattr(value, 'formula'):
            return str(value.formula)
        if isinstance(value, (int, float, str, bool)):
            return value
        return str(value)
    
    # Test with v3 serialize_cell_value
    def v3_serialize(value):
        if value is None:
            return None
        if type(value).__name__ == "ArrayFormula":
            return getattr(value, 'text', repr(value))
        if hasattr(value, 'formula'):
            return str(value.formula)
        if isinstance(value, (int, float, str, bool)):
            return value
        return str(value)
    
    # Test improved serialize_cell_value
    def improved_serialize(value):
        if value is None:
            return None
        if type(value).__name__ == "ArrayFormula":
            # Use the actual formula content, not the object representation
            return str(value.formula) if hasattr(value, 'formula') else str(value)
        if hasattr(value, 'formula'):
            return str(value.formula)
        if isinstance(value, (int, float, str, bool)):
            return value
        return str(value)
    
    print("\nTesting serialization functions:")
    print(f"Current serialize af1: {current_serialize(af1)}")
    print(f"Current serialize af2: {current_serialize(af2)}")
    print(f"Current af1 == af2: {current_serialize(af1) == current_serialize(af2)}")
    print()
    
    print(f"v2 serialize af1: {v2_serialize(af1)}")
    print(f"v2 serialize af2: {v2_serialize(af2)}")
    print(f"v2 af1 == af2: {v2_serialize(af1) == v2_serialize(af2)}")
    print()
    
    print(f"v3 serialize af1: {v3_serialize(af1)}")
    print(f"v3 serialize af2: {v3_serialize(af2)}")
    print(f"v3 af1 == af2: {v3_serialize(af1) == v3_serialize(af2)}")
    print()
    
    print(f"Improved serialize af1: {improved_serialize(af1)}")
    print(f"Improved serialize af2: {improved_serialize(af2)}")
    print(f"Improved af1 == af2: {improved_serialize(af1) == improved_serialize(af2)}")
    print()

if __name__ == "__main__":
    create_array_formula_test()
    test_serialize_comparison()