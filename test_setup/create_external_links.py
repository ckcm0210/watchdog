#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Create Excel files with actual external links for testing
"""

import os
import shutil
from openpyxl import Workbook

def create_source_excel_files():
    """Create source Excel files that will be referenced"""
    
    # Create source file 1
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "SourceData"
    ws1['A1'] = 'Source Value 1'
    ws1['B1'] = 100
    ws1['C1'] = 'Data from File 1'
    
    source1_path = '/home/runner/work/watchdog/watchdog/test_setup/source1.xlsx'
    wb1.save(source1_path)
    
    # Create source file 2
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "DataTable"
    ws2['A1'] = 'Table Data'
    ws2['B1'] = 200
    ws2['C1'] = 'Data from File 2'
    
    source2_path = '/home/runner/work/watchdog/watchdog/test_setup/source2.xlsx'
    wb2.save(source2_path)
    
    print(f"Created source files:")
    print(f"  {source1_path}")
    print(f"  {source2_path}")
    
    return source1_path, source2_path

def create_excel_with_external_links():
    """Create Excel file with external links using Excel application"""
    
    print("Note: To test external links properly, you would need to:")
    print("1. Create Excel files with actual external links using Excel application")
    print("2. Open Excel, create a new workbook")
    print("3. Add formulas like ='[source1.xlsx]SourceData'!A1")
    print("4. Save the file")
    print("5. This will create the proper external link XML structure")
    
    # For now, let's create a simple test file that simulates the structure
    wb = Workbook()
    ws = wb.active
    ws.title = "MainData"
    
    # These won't create actual external links, but will help test the formula resolution
    ws['A1'] = "=[1]SourceData!A1"
    ws['B1'] = "=[2]DataTable!B1"
    ws['C1'] = "=SUM([1]SourceData!B1,[2]DataTable!B1)"
    
    main_path = '/home/runner/work/watchdog/watchdog/test_setup/main_with_external.xlsx'
    wb.save(main_path)
    
    print(f"Created main file with simulated external references: {main_path}")
    return main_path

def create_mock_external_link_xml():
    """Create mock external link XML files for testing"""
    
    # This is a simplified example of what external link XML might look like
    external_link_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<externalLink xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
    <externalBook r:id="rId1">
        <sheetNames>
            <sheetName val="SourceData"/>
        </sheetNames>
    </externalBook>
</externalLink>'''
    
    workbook_rels_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
    <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
    <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink" Target="externalLinks/externalLink1.xml"/>
    <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink" Target="externalLinks/externalLink2.xml"/>
</Relationships>'''
    
    print("Mock external link XML structures created for reference")
    print("In real Excel files, external links are automatically created when you reference other workbooks")

if __name__ == "__main__":
    create_source_excel_files()
    create_excel_with_external_links()
    create_mock_external_link_xml()