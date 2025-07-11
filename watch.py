# -*- coding: utf-8 -*-
"""
Created on Thu Jul 10 16:29:38 2025

@author: kccheng
"""

import os
import time
import csv
import hashlib
import gc
import psutil
import shutil
import tempfile
import gzip
import json
import signal
import threading
from datetime import datetime
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook

# =========== User Config ============
SCAN_ALL_MODE = True

# 🚀 優化選項
USE_LOCAL_CACHE = True
ENABLE_FAST_MODE = True
CACHE_FOLDER = r".\\excel_cache"

# 🔧 變更檢測選項
FORMULA_ONLY_COMPARISON = True  # True=只檢測公式變更, False=檢測公式和值變更
ENABLE_ARRAY_FORMULA_FILTER = True  # 啟用 Array Formula 過濾功能

# 🔧 診斷和恢復選項
ENABLE_TIMEOUT = True          # 啟用超時保護
FILE_TIMEOUT_SECONDS = 120     # 每個檔案最大處理時間 (秒)
ENABLE_MEMORY_MONITOR = True   # 啟用記憶體監控
MEMORY_LIMIT_MB = 2048         # 記憶體限制 (MB)
ENABLE_RESUME = True           # 啟用斷點續傳
RESUME_LOG_FILE = r".\\baseline_progress.log"  # 進度記錄檔

WATCH_FOLDERS = [
    r"\\network_drive\\your_folder1",
    r"\\network_drive\\your_folder2"
]

MANUAL_BASELINE_TARGET = [
    r"\\network_drive\\your_folder1\\somefile.xlsx",
    r"\\network_drive\\your_folder2\\subfolder"
]

LOG_FOLDER = r".\\excel_watch_log"
LOG_FILE_DATE = datetime.now().strftime('%Y%m%d')
CSV_LOG_FILE = os.path.join(LOG_FOLDER, f"excel_change_log_{LOG_FILE_DATE}.csv.gz")
SUPPORTED_EXTS = ('.xlsx', '.xlsm')

MAX_RETRY = 10
RETRY_INTERVAL_SEC = 2
USE_TEMP_COPY = True

WHITELIST_USERS = ['ckcm0210', 'yourwhiteuser']
LOG_WHITELIST_USER_CHANGE = True

FORCE_BASELINE_ON_FIRST_SEEN = [
    r"\\network_drive\\your_folder1\\must_first_baseline.xlsx",
    "force_this_file.xlsx"
]
# =========== End User Config ============

# 全局變數
current_processing_file = None
processing_start_time = None
force_stop = False

def signal_handler(signum, frame):
    """處理 Ctrl+C 中斷"""
    global force_stop
    force_stop = True
    print("\n🛑 收到中斷信號，正在安全停止...")
    if current_processing_file:
        print(f"   目前處理檔案: {current_processing_file}")

signal.signal(signal.SIGINT, signal_handler)

def get_memory_usage():
    """獲取目前記憶體使用量"""
    try:
        process = psutil.Process(os.getpid())
        return process.memory_info().rss / 1024 / 1024  # MB
    except Exception:
        return 0

def check_memory_limit():
    """檢查記憶體是否超限"""
    if not ENABLE_MEMORY_MONITOR:
        return False
    
    current_memory = get_memory_usage()
    if current_memory > MEMORY_LIMIT_MB:
        print(f"⚠️ 記憶體使用量過高: {current_memory:.1f} MB > {MEMORY_LIMIT_MB} MB")
        print("   正在執行垃圾回收...")
        gc.collect()
        new_memory = get_memory_usage()
        print(f"   垃圾回收後: {new_memory:.1f} MB")
        return new_memory > MEMORY_LIMIT_MB
    return False

def save_progress(completed_files, total_files):
    """儲存進度到檔案"""
    if not ENABLE_RESUME:
        return
    
    try:
        progress_data = {
            "timestamp": datetime.now().isoformat(),
            "completed": completed_files,
            "total": total_files,
            "completed_list": completed_files  # 可以改為檔案列表
        }
        
        with open(RESUME_LOG_FILE, 'w', encoding='utf-8') as f:
            json.dump(progress_data, f, ensure_ascii=False, indent=2)
            
    except Exception as e:
        print(f"[WARN] 無法儲存進度: {e}")

def load_progress():
    """載入之前的進度"""
    if not ENABLE_RESUME or not os.path.exists(RESUME_LOG_FILE):
        return None
    
    try:
        with open(RESUME_LOG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"[WARN] 無法載入進度: {e}")
        return None

def timeout_handler():
    """超時處理函數"""
    global current_processing_file, processing_start_time, force_stop
    
    while not force_stop:
        time.sleep(10)  # 每 10 秒檢查一次
        
        if current_processing_file and processing_start_time:
            elapsed = time.time() - processing_start_time
            if elapsed > FILE_TIMEOUT_SECONDS:
                print(f"\n⏰ 檔案處理超時!")
                print(f"   檔案: {current_processing_file}")
                print(f"   已處理時間: {elapsed:.1f} 秒 > {FILE_TIMEOUT_SECONDS} 秒")
                print(f"   將跳過此檔案並繼續...")
                # 這裡可以設置一個標誌來跳過當前檔案
                current_processing_file = None
                processing_start_time = None

def get_all_excel_files(folders):
    all_files = []
    for folder in folders:
        if os.path.isfile(folder):
            if folder.lower().endswith(SUPPORTED_EXTS) and not os.path.basename(folder).startswith('~$'):
                all_files.append(folder)
        elif os.path.isdir(folder):
            for dirpath, _, filenames in os.walk(folder):
                for f in filenames:
                    if f.lower().endswith(SUPPORTED_EXTS) and not f.startswith('~$'):
                        all_files.append(os.path.join(dirpath, f))
    return all_files

def serialize_cell_value(value):
    """快速序列化"""
    if value is None:
        return None
    elif isinstance(value, datetime):
        return value.isoformat()
    elif isinstance(value, (int, float, str, bool)):
        return value
    else:
        return str(value)

def get_cell_formula(cell):
    """
    正確提取 cell 的公式，處理 Array Formula 記憶體地址問題
    """
    if cell.data_type != "f":
        return None
    
    try:
        # 如果是 Array Formula，openpyxl 會返回 ArrayFormula 物件
        # 我們需要提取實際的公式字符串而不是物件的記憶體地址
        formula_value = cell.value
        
        # 處理 Array Formula 物件
        if hasattr(formula_value, 'text'):
            # ArrayFormula 物件有 text 屬性包含實際公式
            formula_str = formula_value.text
        elif hasattr(formula_value, 'ref') and hasattr(formula_value, 'formula'):
            # 某些版本的 openpyxl 使用不同的結構
            formula_str = formula_value.formula
        else:
            # 普通公式或其他情況
            formula_str = str(formula_value)
        
        # 確保公式以 = 開頭
        if formula_str and not formula_str.startswith("="):
            formula_str = "=" + formula_str
            
        return formula_str
        
    except Exception as e:
        # 如果提取失敗，返回字符串形式但去除可能的記憶體地址
        try:
            formula_str = str(cell.value)
            # 移除可能的記憶體地址模式 (如 <ArrayFormula 'formula' (A1:B2)>)
            import re
            # 提取引號中的公式部分
            match = re.search(r"'([^']*)'", formula_str)
            if match:
                formula_str = match.group(1)
            
            if formula_str and not formula_str.startswith("="):
                formula_str = "=" + formula_str
                
            return formula_str
        except:
            return str(cell.value)

def pretty_formula(formula):
    """
    美化公式顯示，去除記憶體地址和其他不必要的資訊
    """
    if not formula:
        return formula
    
    try:
        # 去除可能的記憶體地址模式
        import re
        
        # 模式1: <ArrayFormula 'formula' (range)>
        match = re.search(r"<ArrayFormula\s+'([^']+)'\s+\([^)]+\)>", formula)
        if match:
            return "=" + match.group(1) if not match.group(1).startswith("=") else match.group(1)
        
        # 模式2: <object at 0x...>
        if re.search(r"<.*?at\s+0x[0-9a-fA-F]+>", formula):
            # 如果包含記憶體地址，嘗試提取可能的公式部分
            # 這種情況可能需要從原始 cell 重新提取
            return "[Array Formula - 無法顯示]"
        
        # 普通公式，直接返回
        return formula
        
    except Exception:
        return formula

def filter_array_formula_change(changes):
    """
    過濾 Array Formula 變更，移除僅因記憶體地址變動而產生的假變更
    """
    if not ENABLE_ARRAY_FORMULA_FILTER:
        return changes
    
    filtered_changes = []
    
    for change in changes:
        old_formula = change.get('old_formula')
        new_formula = change.get('new_formula')
        
        # 如果兩個公式都存在，比較它們的實際內容
        if old_formula and new_formula:
            # 使用 pretty_formula 清理公式
            clean_old = pretty_formula(old_formula)
            clean_new = pretty_formula(new_formula)
            
            # 如果清理後的公式相同，則跳過這個變更
            if clean_old == clean_new:
                continue
        
        filtered_changes.append(change)
    
    return filtered_changes

def get_excel_last_author(path):
    try:
        wb = load_workbook(path, read_only=True)
        author = wb.properties.lastModifiedBy
        wb.close()
        return author
    except Exception:
        return None

def copy_to_cache(network_path):
    """🚀 帶診斷的緩存功能"""
    if not USE_LOCAL_CACHE:
        return network_path
    
    try:
        os.makedirs(CACHE_FOLDER, exist_ok=True)
        
        # 檢查原始檔案是否存在和可讀
        if not os.path.exists(network_path):
            raise FileNotFoundError(f"網絡檔案不存在: {network_path}")
        
        if not os.access(network_path, os.R_OK):
            raise PermissionError(f"無法讀取網絡檔案: {network_path}")
        
        file_hash = hashlib.md5(network_path.encode('utf-8')).hexdigest()[:16]
        cache_file = os.path.join(CACHE_FOLDER, f"{file_hash}_{os.path.basename(network_path)}")
        
        # 檢查緩存
        if os.path.exists(cache_file):
            try:
                network_mtime = os.path.getmtime(network_path)
                cache_mtime = os.path.getmtime(cache_file)
                if cache_mtime >= network_mtime:
                    return cache_file
            except Exception:
                pass
        
        # 複製檔案，顯示進度
        network_size = os.path.getsize(network_path)
        print(f"   📥 複製到緩存: {os.path.basename(network_path)} ({network_size/(1024*1024):.1f} MB)")
        
        copy_start = time.time()
        shutil.copy2(network_path, cache_file)
        copy_time = time.time() - copy_start
        
        print(f"      複製完成，耗時 {copy_time:.1f} 秒")
        return cache_file
        
    except Exception as e:
        print(f"   ❌ 緩存失敗: {e}")
        return network_path

def dump_excel_cells_with_timeout(path):
    """🚀 帶超時保護的 Excel 讀取"""
    global current_processing_file, processing_start_time
    
    current_processing_file = path
    processing_start_time = time.time()
    
    try:
        # 檢查檔案基本資訊
        file_size = os.path.getsize(path)
        print(f"   📊 檔案大小: {file_size/(1024*1024):.1f} MB")
        
        # 使用本地緩存
        local_path = copy_to_cache(path)
        
        if ENABLE_FAST_MODE:
            # 快速模式
            print(f"   🚀 使用快速模式讀取...")
            wb = load_workbook(local_path, read_only=True, data_only=False)
            result = {}
            
            worksheet_count = len(wb.worksheets)
            print(f"   📋 工作表數量: {worksheet_count}")
            
            for idx, ws in enumerate(wb.worksheets, 1):
                print(f"      處理工作表 {idx}/{worksheet_count}: {ws.title}")
                
                ws_data = {}
                cell_count = 0
                
                if ws.max_row > 1 or ws.max_column > 1:
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
                                          min_col=1, max_col=ws.max_column):
                        for cell in row:
                            if cell.value is not None:
                                formula = get_cell_formula(cell)
                                
                                ws_data[cell.coordinate] = {
                                    "formula": formula,
                                    "value": serialize_cell_value(cell.value)
                                }
                                cell_count += 1
                
                print(f"         找到 {cell_count} 個有資料的 cell")
                
                if ws_data:
                    result[ws.title] = ws_data
            
            wb.close()
            print(f"   ✅ Excel 讀取完成")
        else:
            # 標準模式
            print(f"   📚 使用標準模式讀取...")
            wb_formula = load_workbook(local_path, data_only=False)
            wb_value = load_workbook(local_path, data_only=True)
            result = {}
            
            for ws_formula, ws_value in zip(wb_formula.worksheets, wb_value.worksheets):
                ws_data = {}
                for row_formula, row_value in zip(ws_formula.iter_rows(), ws_value.iter_rows()):
                    for cell_formula, cell_value in zip(row_formula, row_value):
                        try:
                            formula = get_cell_formula(cell_formula)
                            value = serialize_cell_value(cell_value.value)
                            
                            if formula or (value not in [None, ""]):
                                ws_data[cell_formula.coordinate] = {
                                    "formula": formula,
                                    "value": value
                                }
                        except Exception:
                            pass
                
                if ws_data:
                    result[ws_formula.title] = ws_data
            
            wb_formula.close()
            wb_value.close()
        
        current_processing_file = None
        processing_start_time = None
        return result
        
    except Exception as e:
        current_processing_file = None
        processing_start_time = None
        print(f"   ❌ Excel 讀取失敗: {e}")
        return {}

def hash_excel_content(cells_dict):
    try:
        content_str = json.dumps(cells_dict, sort_keys=True, ensure_ascii=False)
        return hashlib.md5(content_str.encode('utf-8')).hexdigest()
    except Exception:
        return None

def baseline_file_path(base_name):
    return os.path.join(LOG_FOLDER, f"{base_name}.baseline.json.gz")

def save_baseline(baseline_file, data):
    try:
        with gzip.open(baseline_file, 'wt', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, separators=(',', ':'))
    except Exception as e:
        print(f"[ERROR][save_baseline] error saving {baseline_file}: {e}")

def is_force_baseline_file(filepath):
    try:
        lowerfile = filepath.lower()
        for pattern in FORCE_BASELINE_ON_FIRST_SEEN:
            if pattern.lower() in lowerfile:
                return True
        return False
    except Exception:
        return False

def human_readable_size(num_bytes):
    for unit in ['B','KB','MB','GB','TB']:
        if num_bytes < 1024.0:
            return f"{num_bytes:,.2f} {unit}"
        num_bytes /= 1024.0
    return f"{num_bytes:.2f} PB"

def create_baseline_for_files_robust(xlsx_files, skip_force_baseline=True):
    """🛡️ 強化版 baseline 建立，帶診斷和恢復功能"""
    global force_stop
    
    total = len(xlsx_files)
    if total == 0:
        print("[INFO] 沒有需要 baseline 的檔案。")
        return

    print()
    print("=" * 90)
    print(" BASELINE 建立程序 (強化診斷版本) ".center(90, "="))
    print("=" * 90)
    
    # 檢查是否有之前的進度
    progress = load_progress()
    start_index = 0
    if progress and ENABLE_RESUME:
        print(f"🔄 發現之前的進度記錄:")
        print(f"   之前完成: {progress['completed']}/{progress['total']}")
        print(f"   記錄時間: {progress['timestamp']}")
        
        resume = input("是否要從上次中斷的地方繼續? (y/n): ").strip().lower()
        if resume == 'y':
            start_index = progress['completed']
            print(f"   ✅ 從第 {start_index + 1} 個檔案開始")
    
    # 啟動超時監控線程
    if ENABLE_TIMEOUT:
        timeout_thread = threading.Thread(target=timeout_handler, daemon=True)
        timeout_thread.start()
        print(f"⏰ 啟用超時保護: {FILE_TIMEOUT_SECONDS} 秒")
    
    if ENABLE_MEMORY_MONITOR:
        print(f"💾 啟用記憶體監控: {MEMORY_LIMIT_MB} MB")
    
    optimizations = []
    if USE_LOCAL_CACHE:
        optimizations.append("本地緩存")
    if ENABLE_FAST_MODE:
        optimizations.append("快速模式")
    
    print(f"🚀 啟用優化: {', '.join(optimizations)}")
    print(f"📂 Baseline 儲存位置: {os.path.abspath(LOG_FOLDER)}")
    if USE_LOCAL_CACHE:
        print(f"💾 本地緩存位置: {os.path.abspath(CACHE_FOLDER)}")
    print(f"📋 要處理的檔案: {total} 個 Excel (從第 {start_index + 1} 個開始)")
    print(f"⏰ 開始時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    print("-" * 90)
    
    # 確保資料夾存在
    os.makedirs(LOG_FOLDER, exist_ok=True)
    if USE_LOCAL_CACHE:
        os.makedirs(CACHE_FOLDER, exist_ok=True)
    
    baseline_total_size = 0
    success_count = 0
    skip_count = 0
    error_count = 0
    start_time = time.time()
    
    for i in range(start_index, total):
        if force_stop:
            print("\n🛑 收到停止信號，正在安全退出...")
            save_progress(i, total)
            break
            
        file_path = xlsx_files[i]
        base_name = os.path.basename(file_path)
        baseline_file = baseline_file_path(base_name)
        
        # 檢查記憶體
        if check_memory_limit():
            print(f"⚠️ 記憶體使用量過高，暫停 10 秒...")
            time.sleep(10)
            if check_memory_limit():
                print(f"❌ 記憶體仍然過高，停止處理")
                save_progress(i, total)
                break
        
        # 記錄檔案處理時間
        file_start_time = time.time()
        start_time_str = datetime.now().strftime('%H:%M:%S')
        current_memory = get_memory_usage()
        
        print(f"[完成 {i+1:>2}/{total}] [原始#{i+1:>2}] 處理中... (記憶體: {current_memory:.1f}MB)")
        print(f"  檔案: {base_name}")
        
        try:
            # 檢查是否跳過
            if skip_force_baseline and is_force_baseline_file(file_path):
                end_time_str = datetime.now().strftime('%H:%M:%S')
                consumed_time = time.time() - file_start_time
                
                print(f"  結果: [SKIP]")
                print(f"  原因: 屬於 FORCE_BASELINE_ON_FIRST_SEEN")
                print(f"  時間: 從 {start_time_str} 到 {end_time_str} 耗時 {consumed_time:.2f} 秒")
                print()
                
                skip_count += 1
                save_progress(i + 1, total)
                continue
            
            # 🛡️ 使用強化的 Excel 讀取
            cell_data = dump_excel_cells_with_timeout(file_path)
            
            if not cell_data and current_processing_file is None:
                # 可能是超時了
                print(f"  結果: [TIMEOUT]")
                print(f"  原因: 處理超時，跳過此檔案")
                error_count += 1
                save_progress(i + 1, total)
                continue
            
            curr_author = get_excel_last_author(file_path)
            curr_hash = hash_excel_content(cell_data)
            
            # 儲存 baseline
            save_baseline(baseline_file, {
                "last_author": curr_author,
                "content_hash": curr_hash,
                "cells": cell_data
            })
            
            # 計算結果
            size = os.path.getsize(baseline_file)
            baseline_total_size += size
            end_time_str = datetime.now().strftime('%H:%M:%S')
            consumed_time = time.time() - file_start_time
            baseline_name = os.path.basename(baseline_file)
            
            print(f"  結果: [OK]")
            print(f"  Baseline: {baseline_name}")
            print(f"  檔案大小: {human_readable_size(size)} | 累積: {human_readable_size(baseline_total_size)}")
            print(f"  時間: 從 {start_time_str} 到 {end_time_str} 耗時 {consumed_time:.2f} 秒")
            print()
            
            success_count += 1
            save_progress(i + 1, total)
            
        except Exception as e:
            end_time_str = datetime.now().strftime('%H:%M:%S')
            consumed_time = time.time() - file_start_time
            
            print(f"  結果: [ERROR]")
            print(f"  錯誤: {e}")
            print(f"  時間: 從 {start_time_str} 到 {end_time_str} 耗時 {consumed_time:.2f} 秒")
            print()
            
            error_count += 1
            save_progress(i + 1, total)
    
    force_stop = True  # 停止超時監控線程
    
    end_time = time.time()
    total_time = end_time - start_time
    
    print("-" * 90)
    print("🎯 BASELINE 建立完成!")
    print(f"⏱️  總耗時: {total_time:.2f} 秒")
    print(f"✅ 成功: {success_count} 個")
    print(f"⏭️  跳過: {skip_count} 個")
    print(f"❌ 失敗: {error_count} 個")
    print(f"📦 累積 baseline 檔案大小: {human_readable_size(baseline_total_size)}")
    
    if success_count > 0:
        print(f"📊 平均每檔案處理時間: {total_time/total:.2f} 秒")
    
    # 清理進度檔案
    if ENABLE_RESUME and os.path.exists(RESUME_LOG_FILE):
        try:
            os.remove(RESUME_LOG_FILE)
            print(f"🧹 清理進度檔案")
        except Exception:
            pass
    
    print()
    print(f"📁 所有 baseline 檔案存放於: {os.path.abspath(LOG_FOLDER)}")
    if USE_LOCAL_CACHE:
        print(f"💾 本地緩存檔案存放於: {os.path.abspath(CACHE_FOLDER)}")
    print("=" * 90 + "\n")

def compare_excel_changes(file_path):
    """比較 Excel 檔案與 baseline 的變更"""
    try:
        base_name = os.path.basename(file_path)
        baseline_file = baseline_file_path(base_name)
        
        # 載入 baseline
        old_baseline = load_baseline(baseline_file)
        if not old_baseline:
            print(f"[INFO] 沒有 baseline: {base_name}，建立新 baseline...")
            # 建立新 baseline
            cell_data = dump_excel_cells_with_timeout(file_path)
            curr_author = get_excel_last_author(file_path)
            curr_hash = hash_excel_content(cell_data)
            save_baseline(baseline_file, {
                "last_author": curr_author,
                "content_hash": curr_hash,
                "cells": cell_data
            })
            return

        # 讀取現在的檔案
        curr_cells = dump_excel_cells_with_timeout(file_path)
        curr_author = get_excel_last_author(file_path)
        curr_hash = hash_excel_content(curr_cells)
        
        old_cells = old_baseline.get('cells', {})
        old_author = old_baseline.get('last_author', '')
        old_hash = old_baseline.get('content_hash', '')

        # Hash 比較
        if curr_hash == old_hash:
            print(f"[INFO] 檔案無變更: {base_name}")
            return

        print(f"\n🚨 [檔案有變更] {base_name}")
        print(f"  作者: {old_author} → {curr_author}")
        print(f"  Hash: {old_hash[:8]}... → {curr_hash[:8]}...")
        
        # 詳細 cell 比較
        changes = []
        
        # 找出所有 cell 位置
        all_cells = set()
        for ws_name in old_cells.keys():
            all_cells.update([(ws_name, cell) for cell in old_cells[ws_name].keys()])
        for ws_name in curr_cells.keys():
            all_cells.update([(ws_name, cell) for cell in curr_cells[ws_name].keys()])
        
        for ws_name, cell_coord in all_cells:
            old_cell = old_cells.get(ws_name, {}).get(cell_coord, {"formula": None, "value": None})
            curr_cell = curr_cells.get(ws_name, {}).get(cell_coord, {"formula": None, "value": None})
            
            # 根據配置決定比較模式
            if FORMULA_ONLY_COMPARISON:
                # 只比較公式變更
                if old_cell.get('formula') != curr_cell.get('formula'):
                    changes.append({
                        'worksheet': ws_name,
                        'cell': cell_coord,
                        'old_formula': old_cell['formula'],
                        'old_value': old_cell['value'],
                        'new_formula': curr_cell['formula'],
                        'new_value': curr_cell['value'],
                        'change_type': 'formula'
                    })
            else:
                # 比較公式和值變更
                if old_cell != curr_cell:
                    # 判斷變更類型
                    change_type = 'both'
                    if old_cell.get('formula') != curr_cell.get('formula'):
                        change_type = 'formula' if old_cell.get('value') == curr_cell.get('value') else 'both'
                    elif old_cell.get('value') != curr_cell.get('value'):
                        change_type = 'value'
                    
                    changes.append({
                        'worksheet': ws_name,
                        'cell': cell_coord,
                        'old_formula': old_cell['formula'],
                        'old_value': old_cell['value'],
                        'new_formula': curr_cell['formula'],
                        'new_value': curr_cell['value'],
                        'change_type': change_type
                    })
        
        # 過濾 Array Formula 假變更
        changes = filter_array_formula_change(changes)
        
        print_cell_changes_summary(changes)
        
        # 記錄到 CSV
        log_changes_to_csv(file_path, curr_author, changes)
        
        # 更新 baseline
        save_baseline(baseline_file, {
            "last_author": curr_author,
            "content_hash": curr_hash,
            "cells": curr_cells
        })
        
    except Exception as e:
        print(f"[ERROR] 比較檔案失敗: {file_path} - {e}")

def print_cell_changes_summary(changes, max_show=10):
    """🎯 新格式的 cell 變更顯示"""
    try:
        print(f"  變更 cell 數量：{len(changes)}")
        
        # 統計變更類型
        change_types = {}
        for change in changes:
            change_type = change.get('change_type', 'unknown')
            change_types[change_type] = change_types.get(change_type, 0) + 1
        
        if change_types:
            type_summary = ", ".join([f"{k}: {v}" for k, v in change_types.items()])
            print(f"  變更類型統計：{type_summary}")
        
        for i, change in enumerate(changes[:max_show]):
            ws = change['worksheet']
            cell = change['cell']
            old_formula = pretty_formula(change['old_formula']) or ""
            old_value = change['old_value'] or ""
            new_formula = pretty_formula(change['new_formula']) or ""
            new_value = change['new_value'] or ""
            change_type = change.get('change_type', 'unknown')
            
            # 根據變更類型決定顯示內容
            if change_type == 'formula':
                print(f"    [{ws}] {cell} [公式變更]:")
                print(f"        [公式: {old_formula}] -> [公式: {new_formula}]")
                if old_value != new_value:
                    print(f"        [值: {old_value}] -> [值: {new_value}]")
            elif change_type == 'value':
                print(f"    [{ws}] {cell} [值變更]:")
                if old_formula:
                    print(f"        [公式: {old_formula}] (未變更)")
                print(f"        [值: {old_value}] -> [值: {new_value}]")
            else:
                # 檢查公式長度決定格式
                formula_line = f"[公式: {old_formula}] -> [公式: {new_formula}]"
                value_line = f"[值: {old_value}] -> [值: {new_value}]"
                
                # 如果公式行太長（超過 80 字符），就分行顯示
                if len(formula_line) > 80:
                    print(f"    [{ws}] {cell}:")
                    print(f"        [公式: {old_formula}]")
                    print(f"        -> [公式: {new_formula}]")
                    print(f"        {value_line}")
                else:
                    print(f"    [{ws}] {cell}:")
                    print(f"        {formula_line}")
                    print(f"        {value_line}")
        
        if len(changes) > max_show:
            print(f"    ... 其餘 {len(changes) - max_show} 個 cell 省略 ...")
    except Exception as e:
        print(f"[ERROR][print_cell_changes_summary] {e}")

def log_changes_to_csv(file_path, author, changes):
    """記錄變更到 CSV"""
    try:
        os.makedirs(LOG_FOLDER, exist_ok=True)
        
        with gzip.open(CSV_LOG_FILE, 'at', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            for change in changes:
                writer.writerow([
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    file_path,
                    author,
                    change['worksheet'],
                    change['cell'],
                    pretty_formula(change['old_formula']),
                    change['old_value'],
                    pretty_formula(change['new_formula']),
                    change['new_value'],
                    change.get('change_type', 'unknown')
                ])
    except Exception as e:
        print(f"[ERROR] 記錄 CSV 失敗: {e}")

def load_baseline(baseline_file):
    """載入 baseline 檔案"""
    try:
        if os.path.exists(baseline_file):
            with gzip.open(baseline_file, 'rt', encoding='utf-8') as f:
                return json.load(f)
        return None
    except Exception as e:
        print(f"[ERROR][load_baseline] {baseline_file}: {e}")
        return None

def print_console_header():
    print("\n" + "="*80)
    print(" Excel File Change Watcher (診斷強化版本) ".center(80, "-"))
    print("="*80 + "\n")

def start_watchdog_monitor():
    """啟動 Watchdog 監控"""
    print("\n" + "=" * 80)
    print(" 啟動 Excel 檔案監控 ".center(80, "="))
    print("=" * 80)
    
    print("  監控資料夾:")
    for folder in WATCH_FOLDERS:
        print(f"    📂 {folder}")
    
    print(f"  支援檔案: {SUPPORTED_EXTS}")
    print(f"  變更記錄: {CSV_LOG_FILE}")
    print()
    
    event_handler = ExcelChangeHandler()
    observer = Observer()
    
    for folder in WATCH_FOLDERS:
        if os.path.exists(folder):
            observer.schedule(event_handler, folder, recursive=True)
            print(f"✅ 已監控: {folder}")
        else:
            print(f"❌ 資料夾不存在: {folder}")
    
    print("\n🔍 監控中... (按 Ctrl+C 停止)")
    print("-" * 80)
    
    try:
        observer.start()
        while not force_stop:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\n🛑 收到停止信號...")
    finally:
        observer.stop()
        observer.join()
        print("📄 監控已停止")

class ExcelChangeHandler(FileSystemEventHandler):
    def __init__(self):
        self.processing_files = set()
        
    def on_modified(self, event):
        if event.is_directory:
            return
            
        file_path = event.src_path
        if not file_path.lower().endswith(SUPPORTED_EXTS):
            return
            
        filename = os.path.basename(file_path)
        if filename.startswith('~$'):
            return
            
        # 避免重複處理同一檔案
        if file_path in self.processing_files:
            return
            
        self.processing_files.add(file_path)
        
        try:
            # 等待檔案寫入完成
            time.sleep(2)
            
            print(f"\n📝 [檔案修改事件] {filename}")
            print(f"   完整路徑: {file_path}")
            print(f"   時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            # 比較變更
            compare_excel_changes(file_path)
            
        except Exception as e:
            print(f"[ERROR] 處理檔案事件失敗: {file_path} - {e}")
        finally:
            self.processing_files.discard(file_path)
# ============= 其他函數保持原樣... ============

if __name__ == "__main__":
    try:
        print_console_header()
        print("  監控資料夾:")
        for folder in WATCH_FOLDERS:
            print(f"    - {folder}")
        print(f"  支援副檔名: {SUPPORTED_EXTS}")
        print(f"  目前使用者: {os.getlogin()}")  # 應該顯示 ckcm0210
        
        optimizations = []
        if USE_LOCAL_CACHE:
            optimizations.append("本地緩存")
        if ENABLE_FAST_MODE:
            optimizations.append("快速模式")
        if ENABLE_TIMEOUT:
            optimizations.append(f"超時保護({FILE_TIMEOUT_SECONDS}s)")
        if ENABLE_MEMORY_MONITOR:
            optimizations.append(f"記憶體監控({MEMORY_LIMIT_MB}MB)")
        if ENABLE_RESUME:
            optimizations.append("斷點續傳")
        if FORMULA_ONLY_COMPARISON:
            optimizations.append("公式專用比較")
        if ENABLE_ARRAY_FORMULA_FILTER:
            optimizations.append("Array公式過濾")
        
        print(f"  🚀 啟用功能: {', '.join(optimizations)}")
        print(f"  📂 Baseline 儲存位置: {os.path.abspath(LOG_FOLDER)}")
        if USE_LOCAL_CACHE:
            print(f"  💾 本地緩存位置: {os.path.abspath(CACHE_FOLDER)}")
        
        # 確保資料夾存在
        os.makedirs(LOG_FOLDER, exist_ok=True)
        if USE_LOCAL_CACHE:
            os.makedirs(CACHE_FOLDER, exist_ok=True)

        if SCAN_ALL_MODE:
            all_files = get_all_excel_files(WATCH_FOLDERS)
            print(f"總共 find 到 {len(all_files)} 個 Excel file.")
            create_baseline_for_files_robust(all_files, skip_force_baseline=True)
            print("baseline scan 完成！\n")
        else:
            target_files = get_all_excel_files(MANUAL_BASELINE_TARGET)
            print(f"手動指定 baseline，合共 {len(target_files)} 個 Excel file.")
            create_baseline_for_files_robust(target_files, skip_force_baseline=False)
            print("手動 baseline 完成！\n")

        # 🚀 啟動 Watchdog 監控
        start_watchdog_monitor()
        
    except Exception as e:
        print(f"[ERROR][main] 程式主流程 error: {e}")
        import traceback
        traceback.print_exc()
