#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Enhanced watchdog with ArrayFormula filtering and external link mapping
"""

import os
import time
import csv
import hashlib
import gc
import psutil
import shutil
import tempfile
import gzip
import json
import signal
import threading
import zipfile
import xml.etree.ElementTree as ET
import re
from datetime import datetime
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook

# Global variables from original code
current_processing_file = None
processing_start_time = None
force_stop = False

def serialize_cell_value(value):
    """
    Enhanced serialize function that handles ArrayFormula objects properly
    by comparing their formula content rather than object address
    """
    if value is None:
        return None
    
    # Handle ArrayFormula objects - the main fix for requirement 1
    if type(value).__name__ == "ArrayFormula":
        # Get the actual formula content, not the object representation
        if hasattr(value, 'text'):
            return str(value.text)
        elif hasattr(value, 'formula'):
            return str(value.formula)
        else:
            # This shouldn't happen but provide a fallback
            return str(value)
    
    # Handle other objects with formula attribute
    if hasattr(value, 'formula'):
        return str(value.formula)
    
    # Handle standard types
    if isinstance(value, (int, float, str, bool)):
        return value
    
    # Handle datetime
    if isinstance(value, datetime):
        return value.isoformat()
    
    # Default to string representation
    return str(value)

def extract_external_links(excel_file_path):
    """
    Extract external link mappings from Excel file
    Returns a dictionary mapping [n] indices to file paths
    """
    external_link_mapping = {}
    
    try:
        with zipfile.ZipFile(excel_file_path, 'r') as zip_ref:
            # First, find external link files
            external_link_files = [name for name in zip_ref.namelist() 
                                 if 'externalLink' in name.lower() and name.endswith('.xml')]
            
            if not external_link_files:
                return external_link_mapping
            
            # Read workbook relationships to get external link mappings
            workbook_rels_path = None
            for name in zip_ref.namelist():
                if name.endswith('workbook.xml.rels'):
                    workbook_rels_path = name
                    break
            
            if not workbook_rels_path:
                return external_link_mapping
            
            # Parse workbook relationships
            rels_content = zip_ref.read(workbook_rels_path)
            rels_root = ET.fromstring(rels_content)
            
            # Find external link relationships
            external_link_rels = {}
            for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                rel_type = rel.get('Type', '')
                if 'externalLink' in rel_type:
                    rel_id = rel.get('Id', '')
                    target = rel.get('Target', '')
                    external_link_rels[rel_id] = target
            
            # Parse each external link file
            for ext_link_file in external_link_files:
                try:
                    ext_link_content = zip_ref.read(ext_link_file)
                    ext_link_root = ET.fromstring(ext_link_content)
                    
                    # Find external book references
                    for ext_book in ext_link_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}externalBook'):
                        # Get the relationship ID
                        rel_id = ext_book.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id', '')
                        
                        if rel_id in external_link_rels:
                            target_path = external_link_rels[rel_id]
                            
                            # Extract the index from the filename or use a counter
                            # This is a simplified approach - in real Excel files, the mapping might be different
                            match = re.search(r'externalLink(\d+)', ext_link_file)
                            if match:
                                index = int(match.group(1))
                                external_link_mapping[index] = target_path
                                
                except Exception as e:
                    print(f"[DEBUG] Error parsing external link file {ext_link_file}: {e}")
                    continue
            
    except Exception as e:
        print(f"[DEBUG] Error extracting external links from {excel_file_path}: {e}")
    
    return external_link_mapping

def resolve_external_references(formula, external_link_mapping):
    """
    Resolve [n]Table! references in formulas using the external link mapping
    """
    if not formula or not external_link_mapping:
        return formula
    
    # Pattern to match [n]Table! or [n]Sheet! references
    pattern = r'\[(\d+)\]([^!]+)!'
    
    def replace_ref(match):
        index = int(match.group(1))
        sheet_name = match.group(2)
        
        if index in external_link_mapping:
            file_path = external_link_mapping[index]
            # Extract filename from path for display
            filename = os.path.basename(file_path) if file_path else f"ExternalFile{index}"
            return f"[{filename}]{sheet_name}!"
        else:
            return match.group(0)  # Return original if not found
    
    resolved_formula = re.sub(pattern, replace_ref, formula)
    return resolved_formula

def dump_excel_cells_with_timeout(path, show_sheet_detail=True, silent=False):
    """
    Enhanced Excel reading with external link mapping support
    """
    global current_processing_file, processing_start_time
    
    current_processing_file = path
    processing_start_time = time.time()
    
    try:
        file_size = os.path.getsize(path)
        if not silent:
            print(f"   📊 檔案大小: {file_size/(1024*1024):.1f} MB")
        
        # Extract external link mappings
        external_link_mapping = extract_external_links(path)
        if external_link_mapping and not silent:
            print(f"   🔗 發現外部連結映射: {external_link_mapping}")
        
        # Load workbook
        wb = load_workbook(path, read_only=True, data_only=False)
        result = {}
        worksheet_count = len(wb.worksheets)
        
        if not silent:
            print(f"   📋 工作表數量: {worksheet_count}")
        
        for idx, ws in enumerate(wb.worksheets, 1):
            cell_count = 0
            ws_data = {}
            
            if ws.max_row > 1 or ws.max_column > 1:
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
                                      min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if cell.value is not None:
                            formula = None
                            if cell.data_type == "f":
                                formula = str(cell.value)
                                if not formula.startswith("="):
                                    formula = "=" + formula
                                
                                # Resolve external references in formula
                                if external_link_mapping:
                                    resolved_formula = resolve_external_references(formula, external_link_mapping)
                                    if resolved_formula != formula:
                                        formula = resolved_formula
                            
                            ws_data[cell.coordinate] = {
                                "formula": formula,
                                "value": serialize_cell_value(cell.value)
                            }
                            cell_count += 1
            
            if show_sheet_detail and not silent:
                print(f"      處理工作表 {idx}/{worksheet_count}: {ws.title}（{cell_count} 有資料 cell）")
            
            if ws_data:
                result[ws.title] = ws_data
        
        wb.close()
        
        if not silent:
            print(f"   ✅ Excel 讀取完成")
        
        return result
        
    except Exception as e:
        if not silent:
            print(f"   ❌ Excel 讀取失敗: {e}")
        return {}
    finally:
        current_processing_file = None
        processing_start_time = None

def print_cell_changes_summary(changes, max_show=10):
    """
    Enhanced cell changes summary with external link information
    """
    try:
        print(f"  變更 cell 數量：{len(changes)}")
        maxlen = 50
        
        for i, change in enumerate(changes[:max_show]):
            ws, cell = change['worksheet'], change['cell']
            old_f, old_v = change['old_formula'] or "", change['old_value'] or ""
            new_f, new_v = change['new_formula'] or "", change['new_value'] or ""
            
            print(f"    [{ws}] {cell}:")
            
            # Handle formula changes
            if old_f != new_f:
                # Check if this is just an ArrayFormula object address change
                if (('[openpyxl.worksheet.formula.ArrayFormula object at' in str(old_f) or
                     '[openpyxl.worksheet.formula.ArrayFormula object at' in str(new_f)) and
                    old_v == new_v):
                    print(f"        [公式] ArrayFormula 物件地址變更 (內容相同) - 已過濾")
                else:
                    if len(str(old_f)) > maxlen or len(str(new_f)) > maxlen:
                        print(f"        [公式] '{old_f}'\n              -> '{new_f}'")
                    else:
                        print(f"        [公式] '{old_f}' -> '{new_f}'")
            
            # Handle value changes
            if old_v != new_v:
                if len(str(old_v)) > maxlen or len(str(new_v)) > maxlen:
                    print(f"        [值]   '{old_v}'\n              -> '{new_v}'")
                else:
                    print(f"        [值]   '{old_v}' -> '{new_v}'")
        
        if len(changes) > max_show:
            print(f"    ... 其餘 {len(changes) - max_show} 個 cell 省略 ...")
            
    except Exception as e:
        print(f"[ERROR][print_cell_changes_summary] {e}")

# Test function
def test_enhancements():
    """Test the enhanced functionality"""
    print("=" * 60)
    print("Testing Enhanced Watchdog Functionality")
    print("=" * 60)
    
    # Test ArrayFormula handling
    from openpyxl.worksheet.formula import ArrayFormula
    
    af1 = ArrayFormula(ref="A1:A3")
    af1.text = "SUM(A1:A3)"
    
    af2 = ArrayFormula(ref="A1:A3")
    af2.text = "SUM(A1:A3)"
    
    print("ArrayFormula Test:")
    print(f"af1 serialized: {serialize_cell_value(af1)}")
    print(f"af2 serialized: {serialize_cell_value(af2)}")
    print(f"Serialized values equal: {serialize_cell_value(af1) == serialize_cell_value(af2)}")
    
    # Test external link mapping
    test_file = '/home/runner/work/watchdog/watchdog/test_setup/test_external_links.xlsx'
    if os.path.exists(test_file):
        print(f"\nExternal Link Mapping Test:")
        mapping = extract_external_links(test_file)
        print(f"External links found: {mapping}")
        
        # Test formula resolution
        test_formulas = [
            "=[1]Sheet1!A1",
            "=[2]Data!B1",
            "=[3]Table!C1",
            "=SUM([1]Sheet1!A1:[1]Sheet1!A10)"
        ]
        
        for formula in test_formulas:
            resolved = resolve_external_references(formula, mapping)
            print(f"Original: {formula}")
            print(f"Resolved: {resolved}")
            print()

if __name__ == "__main__":
    test_enhancements()